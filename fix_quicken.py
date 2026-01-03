#!/usr/bin/env python3
"""
quicken_downfill_dates.py

Per worksheet:
1) Delete the top four rows
2) Find the row where Column B equals exactly "TOTAL INFLOWS" (case-sensitive)
   - If found: move back two rows (row_index - 2) and delete that row and all rows below it
3) Detect the Date column from the header row ("Date", case-insensitive)
4) If found, down-fill blank Date cells with the most recent valid date above

Input:  -i / --input  path/to/file.xlsx
Output: path/to/file-clean.xlsx  (adds "-clean" before the extension)
"""

import argparse
import os
from datetime import date, datetime
from typing import Optional

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell


ROWS_TO_DELETE_TOP = 4
TOTAL_INFLOWS_TEXT = "TOTAL INFLOWS"
TOTAL_INFLOWS_COL = 2  # Column B


def is_real_date_cell(cell: Cell) -> bool:
    """True if the cell contains a usable Excel date."""
    v = cell.value
    if v is None:
        return False
    if isinstance(v, (datetime, date)):
        return True
    # Excel date serial with date-like number format
    if isinstance(v, (int, float)) and getattr(cell, "is_date", False):
        return True
    return False


def is_blank(cell: Cell) -> bool:
    v = cell.value
    if v is None:
        return True
    if isinstance(v, str) and v.strip() == "":
        return True
    return False


def delete_top_rows(ws, n: int) -> None:
    """Delete the top n rows of the worksheet."""
    if ws.max_row >= 1:
        ws.delete_rows(1, min(n, ws.max_row))


def trim_after_total_inflows(ws) -> int:
    """
    After top rows are deleted, locate "TOTAL INFLOWS" in Column B.
    If found at row r, compute start_delete = r - 2, then delete from start_delete to end.

    Returns number of rows deleted (0 if not found / nothing deleted).
    """
    max_row = ws.max_row
    found_row = None

    # Scan Column B only for speed and clarity
    for r in range(1, max_row + 1):
        v = ws.cell(row=r, column=TOTAL_INFLOWS_COL).value
        if v == TOTAL_INFLOWS_TEXT:  # exact, case-sensitive match
            found_row = r
            break

    if found_row is None:
        return 0

    start_delete = found_row - 2
    if start_delete < 1:
        start_delete = 1

    # Recompute max_row just in case (though it shouldn't change during scan)
    max_row = ws.max_row
    rows_to_delete = max_row - start_delete + 1
    if rows_to_delete > 0:
        ws.delete_rows(start_delete, rows_to_delete)
        return rows_to_delete

    return 0


def find_header_column_index(ws, header_name: str) -> Optional[int]:
    """
    Find a column by scanning the first row (header row).
    """
    first_row = next(ws.iter_rows(min_row=1, max_row=1), [])
    for c in first_row:
        if isinstance(c.value, str) and c.value.strip().lower() == header_name.lower():
            return c.column  # 1-based
    return None


def downfill_dates_in_worksheet(ws) -> int:
    """Down-fill Date column for a worksheet. Returns number of filled cells."""
    date_col = find_header_column_index(ws, "Date")
    if date_col is None:
        return 0
    updated = 0

    last_date_value: Optional[object] = None
    last_date_number_format: Optional[str] = None

    # Start after header row
    for row in ws.iter_rows(min_row=2):
        cell = row[date_col - 1]

        if is_real_date_cell(cell):
            last_date_value = cell.value
            if cell.number_format:
                last_date_number_format = cell.number_format
            continue

        if is_blank(cell) and last_date_value is not None:
            cell.value = last_date_value
            if last_date_number_format:
                cell.number_format = last_date_number_format
            updated += 1

    return updated


def downfill_description_in_worksheet(ws) -> int:
    """Down-fill Description column for a worksheet. Returns number of filled cells."""
    desc_col = find_header_column_index(ws, "Description")
    if desc_col is None:
        return 0
    updated = 0

    last_desc_value: Optional[object] = None

    # Start after header row
    for row in ws.iter_rows(min_row=2):
        cell = row[desc_col - 1]

        if not is_blank(cell):
            last_desc_value = cell.value
            continue

        if last_desc_value is not None:
            cell.value = last_desc_value
            updated += 1

    return updated


def output_filename(input_path: str) -> str:
    base, ext = os.path.splitext(input_path)
    if ext.lower() != ".xlsx":
        ext = ".xlsx"
    return f"{base}-clean{ext}"


def main():
    parser = argparse.ArgumentParser(
        description="Clean Quicken XLSX: drop top rows, trim after TOTAL INFLOWS, and down-fill dates when a Date header is present."
    )
    parser.add_argument("-i", "--input", required=True, help="Input XLSX file")
    args = parser.parse_args()

    in_path = args.input
    out_path = output_filename(in_path)

    wb = load_workbook(in_path, data_only=False)

    total_filled = 0
    total_trimmed_rows = 0
    total_description_filled = 0

    for ws in wb.worksheets:
        delete_top_rows(ws, ROWS_TO_DELETE_TOP)
        total_trimmed_rows += trim_after_total_inflows(ws)
        total_filled += downfill_dates_in_worksheet(ws)
        total_description_filled += downfill_description_in_worksheet(ws)

    wb.save(out_path)

    print(f"Saved cleaned file: {out_path}")
    print(f"Deleted top {ROWS_TO_DELETE_TOP} rows per worksheet")
    print(f"Trimmed {total_trimmed_rows} rows total after '{TOTAL_INFLOWS_TEXT}' rule")
    print(f"Filled {total_filled} blank date cells across {len(wb.worksheets)} worksheet(s)")
    print(
        f"Filled {total_description_filled} blank description cells across {len(wb.worksheets)} worksheet(s)"
    )


if __name__ == "__main__":
    main()
    
